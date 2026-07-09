"""
Branch Inventory Dashboard
===========================
Reads a monthly inventory workbook (e.g. "CCK-TABLET", "CCK-NICHE",
"LST-TABLE & NICHE", "TLT-TABLE & NICHE") and produces a clean per-sheet
breakdown of Total / Sold / Balance / Balance Value, correctly excluding
sub-total, total, and repeated-header rows.

Design notes (why it's built this way):
- Sheets have MERGED cells for the identifier column (BLOCK/PRODUCT) and
  sometimes NAME — continuation rows show up as blank/None in that column.
  Forward-filling those columns is required, otherwise you either drop
  legitimate data (undercount) or crash the aggregation.
- Summary/sub-total/footer rows are removed by scanning EVERY cell in a row
  for the literal words TOTAL / SUB / SUM. This is what actually filters
  out rows like "BLK A NICHE SUB TOTAL:", "TABLET TOTAL:", "NICHE TOTAL:",
  " ALL PRODUCT TOTAL:", AND the repeated header row that sometimes appears
  mid-sheet (e.g. a stray "BLOCK | LAUNCH DATE | ... | TOTAL" row) — that
  row also contains the word TOTAL, so it's caught by the same rule.
- LOT TYPE classification uses an EXACT dictionary lookup (after
  strip+upper), never substring/keyword matching. This is the fix for the
  10,098 vs 10,070 bug: a keyword rule like `"SG" in lot_type` will also
  match "TWIN SG" (a completely different lot type) since it contains "SG"
  as a substring. Exact matching cannot make that mistake.
- Sheets are handled generically by locating columns via their header TEXT
  (not fixed positions), so the same code works whether the sheet has a
  "BLOCK" or "PRODUCT" identifier column, and whether or not it has a
  "LOT TYPE" column at all (pure tablet sheets don't).
"""

import io
from collections import defaultdict

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Branch Inventory Dashboard", layout="wide")

# =============================================================================
# CONFIG
# =============================================================================

# Any row containing one of these words (in ANY cell) is treated as a
# summary / sub-total / footer / repeated-header row and dropped.
FOOTER_KEYWORDS = ["TOTAL", "SUB", "SUM"]

# Strict, exact-match classification. Keys must be UPPERCASE + stripped.
# Add new lot type codes here as the reporting team introduces them —
# do NOT switch this to substring matching.
LOT_TYPE_MAP = {
    "SINGLE": "Single",
    "SG": "Single",
    "DOUBLE": "Double",
    "DB": "Double",
    "FAMILY": "Family",
    "TOWER": "Tower",
    "BUDDHA": "Buddha",
    "U-BUDDHA": "Buddha",
    "TWIN DB": "Special",
    "TWIN SG": "Special",
    "M-TWS": "Special",
    "M-TWD": "Special",
}

CATEGORY_ORDER = [
    "Niche - Single",
    "Niche - Double",
    "Niche - Family",
    "Niche - Buddha",
    "Niche - Tower",
    "Niche - Special",
    "Tablet",
]


# =============================================================================
# PARSING HELPERS
# =============================================================================

def classify_lot_type(raw) -> str:
    """Exact dictionary classification. Unknown codes are surfaced (not
    silently dropped, not silently miscounted) so they can be reviewed and
    added to LOT_TYPE_MAP."""
    if raw is None or str(raw).strip() == "":
        return "Unclassified (BLANK)"
    key = str(raw).strip().upper()
    return LOT_TYPE_MAP.get(key, f"Unclassified ({key})")


def try_parse_number(val):
    """Robustly convert a cell to float. Returns None if the cell doesn't
    represent a real number (blank, '-', 'NIL', '#DIV/0!', text notes, etc.)
    so callers can distinguish 'no data' from 'zero'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "")
    if s in ("", "-", "\u2014", "N/A", "NIL", "#DIV/0!"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def to_number(val) -> float:
    """Like try_parse_number but always returns a usable float (0.0 for
    non-numeric/missing) — used once we've already decided a row is valid
    data and just need safe arithmetic."""
    parsed = try_parse_number(val)
    return parsed if parsed is not None else 0.0


def find_header_row(ws, max_scan: int = 10):
    """Locate the header row by looking for a cell literally equal to
    'BLOCK' or 'PRODUCT' (the two identifier-column headers used across
    these report formats)."""
    for r in range(1, max_scan + 1):
        row_vals = [c.value for c in ws[r]]
        for v in row_vals:
            if isinstance(v, str) and v.strip().upper() in ("BLOCK", "PRODUCT"):
                return r, row_vals
    return None, None


def find_col_exact(header_vals, name: str):
    for i, h in enumerate(header_vals):
        if isinstance(h, str) and h.strip().upper() == name:
            return i
    return None


def find_col_contains(header_vals, substr: str, exclude=None):
    exclude = exclude or []
    for i, h in enumerate(header_vals):
        if not isinstance(h, str):
            continue
        hu = h.strip().upper()
        if substr in hu and not any(ex in hu for ex in exclude):
            return i
    return None


def is_footer_or_summary_row(row_vals) -> bool:
    for v in row_vals:
        if isinstance(v, str) and any(kw in v.strip().upper() for kw in FOOTER_KEYWORDS):
            return True
    return False


def parse_sheet(ws):
    """Parse one worksheet into a list of clean row-level records.

    Returns (records, header_vals). records is [] if the sheet doesn't
    look like an inventory sheet at all (no BLOCK/PRODUCT header found).
    """
    header_row_idx, header_vals = find_header_row(ws)
    if header_row_idx is None:
        return [], None

    idx_id = find_col_exact(header_vals, "BLOCK")
    if idx_id is None:
        idx_id = find_col_exact(header_vals, "PRODUCT")
    idx_lot = find_col_exact(header_vals, "LOT TYPE")
    idx_total = find_col_exact(header_vals, "TOTAL")
    idx_sold = find_col_exact(header_vals, "TOTAL SOLD")
    idx_bal = find_col_exact(header_vals, "BALANCE")
    idx_balval = find_col_contains(header_vals, "BALANCE $", exclude=["NETT"])

    if idx_id is None or idx_total is None:
        return [], header_vals

    id_header_name = header_vals[idx_id].strip().upper()

    records = []
    last_id = None
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]

        if all(v is None for v in row_vals):
            continue  # fully blank spacer row

        if is_footer_or_summary_row(row_vals):
            continue  # sub-total / total / repeated-header row

        # Forward-fill the merged identifier column (BLOCK/PRODUCT)
        raw_id = row_vals[idx_id] if idx_id < len(row_vals) else None
        if raw_id is not None and str(raw_id).strip() != "":
            last_id = raw_id
        record_id = last_id

        # A row only counts as real DATA if it has a numeric TOTAL.
        # This drops explanatory footnote rows and empty separator rows
        # that survived the blank-row check (e.g. a row with only a
        # LAUNCH DATE and nothing else).
        total_num = try_parse_number(row_vals[idx_total]) if idx_total < len(row_vals) else None
        if total_num is None:
            continue

        lot_raw = row_vals[idx_lot] if (idx_lot is not None and idx_lot < len(row_vals)) else None

        # Classification: does this sheet distinguish TABLET vs NICHE via
        # the identifier column (LST / TLT style, header = "PRODUCT"), or
        # is the whole sheet one product type (CCK style, header = "BLOCK")?
        if idx_lot is not None:
            if id_header_name == "PRODUCT":
                product_val = str(record_id).strip().upper() if record_id else ""
                if "TABLET" in product_val:
                    category = "Tablet"
                else:
                    category = f"Niche - {classify_lot_type(lot_raw)}"
            else:
                category = f"Niche - {classify_lot_type(lot_raw)}"
        else:
            category = "Tablet"  # sheet has no LOT TYPE column at all

        bal_num = try_parse_number(row_vals[idx_bal]) if idx_bal is not None and idx_bal < len(row_vals) else None
        sold_num = try_parse_number(row_vals[idx_sold]) if idx_sold is not None and idx_sold < len(row_vals) else None
        balval_num = try_parse_number(row_vals[idx_balval]) if idx_balval is not None and idx_balval < len(row_vals) else None

        bal_clean = bal_num if bal_num is not None else 0.0
        # Fall back to Total - Balance if there's no explicit "TOTAL SOLD" column
        sold_clean = sold_num if sold_num is not None else (total_num - bal_clean)

        records.append({
            "Row": r,
            "Identifier": record_id,
            "LotTypeRaw": lot_raw,
            "Category": category,
            "Total": total_num,
            "Sold": sold_clean,
            "Balance": bal_clean,
            "BalanceValue": (balval_num or 0.0) * 1000,  # sheet stores $ in '000s
        })

    return records, header_vals


def aggregate(records):
    agg = defaultdict(lambda: {"Total": 0.0, "Sold": 0.0, "Balance": 0.0, "BalanceValue": 0.0, "Rows": 0})
    for rec in records:
        a = agg[rec["Category"]]
        a["Total"] += rec["Total"]
        a["Sold"] += rec["Sold"]
        a["Balance"] += rec["Balance"]
        a["BalanceValue"] += rec["BalanceValue"]
        a["Rows"] += 1
    return agg


def agg_to_dataframe(agg: dict) -> pd.DataFrame:
    rows = []
    for category, v in agg.items():
        rows.append({
            "Category": category,
            "Total": v["Total"],
            "Balance": v["Balance"],
            "Sold": v["Sold"],
            "Balance %": (v["Balance"] / v["Total"] * 100) if v["Total"] else 0.0,
            "Sold %": (v["Sold"] / v["Total"] * 100) if v["Total"] else 0.0,
            "Value of Balance": v["BalanceValue"],
            "Source Rows": v["Rows"],
        })
    df = pd.DataFrame(rows)
    if df.empty:
        return df

    # Sort using the preferred display order, unknown categories go last
    df["_sort"] = df["Category"].apply(
        lambda c: CATEGORY_ORDER.index(c) if c in CATEGORY_ORDER else len(CATEGORY_ORDER)
    )
    df = df.sort_values(["_sort", "Category"]).drop(columns="_sort").reset_index(drop=True)

    # Append a Total row
    total_row = {
        "Category": "Total",
        "Total": df["Total"].sum(),
        "Balance": df["Balance"].sum(),
        "Sold": df["Sold"].sum(),
        "Balance %": (df["Balance"].sum() / df["Total"].sum() * 100) if df["Total"].sum() else 0.0,
        "Sold %": (df["Sold"].sum() / df["Total"].sum() * 100) if df["Total"].sum() else 0.0,
        "Value of Balance": df["Value of Balance"].sum(),
        "Source Rows": df["Source Rows"].sum(),
    }
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    return df


def format_display(df: pd.DataFrame) -> pd.DataFrame:
    """String-format numeric columns for display in st.dataframe."""
    if df.empty:
        return df
    out = df.copy()
    out["Total"] = out["Total"].apply(lambda x: f"{x:,.0f}")
    out["Balance"] = out["Balance"].apply(lambda x: f"{x:,.0f}")
    out["Sold"] = out["Sold"].apply(lambda x: f"{x:,.0f}")
    out["Balance %"] = out["Balance %"].apply(lambda x: f"{x:.2f}%")
    out["Sold %"] = out["Sold %"].apply(lambda x: f"{x:.2f}%")
    out["Value of Balance"] = out["Value of Balance"].apply(lambda x: f"${x:,.0f}")
    out["Source Rows"] = out["Source Rows"].apply(lambda x: f"{x:,.0f}")
    return out


# =============================================================================
# STREAMLIT APP
# =============================================================================

st.title("🏢 Branch Inventory Dashboard")
st.caption("Upload the monthly inventory workbook. Every sheet is parsed generically — "
           "no sheet names or row numbers are hard-coded.")

uploaded_file = st.file_uploader("Upload inventory Excel file (.xlsx)", type=["xlsx"])

debug_mode = st.sidebar.checkbox("🔍 Debug mode (show row-level detail per category)", value=False)

if uploaded_file is not None:
    wb = load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)

    all_records_by_sheet = {}
    unclassified_found = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        records, header_vals = parse_sheet(ws)
        if not records:
            continue
        all_records_by_sheet[sheet_name] = records
        for rec in records:
            if "Unclassified" in rec["Category"]:
                unclassified_found.add((sheet_name, rec["Category"], rec["Row"]))

    if not all_records_by_sheet:
        st.error("No recognizable inventory sheets found (expected a header row containing "
                  "'BLOCK' or 'PRODUCT').")
        st.stop()

    if unclassified_found:
        with st.expander(f"⚠️ {len(unclassified_found)} row(s) had an unrecognized LOT TYPE — "
                          f"they were kept but bucketed separately, not silently dropped or "
                          f"miscounted. Add them to LOT_TYPE_MAP to classify properly.", expanded=False):
            udf = pd.DataFrame(sorted(unclassified_found), columns=["Sheet", "Category", "Excel Row"])
            st.dataframe(udf, use_container_width=True, hide_index=True)

    tab_names = list(all_records_by_sheet.keys()) + ["📊 Combined Summary"]
    tabs = st.tabs(tab_names)

    combined_agg = defaultdict(lambda: {"Total": 0.0, "Sold": 0.0, "Balance": 0.0, "BalanceValue": 0.0, "Rows": 0})

    for tab, sheet_name in zip(tabs, all_records_by_sheet.keys()):
        with tab:
            records = all_records_by_sheet[sheet_name]
            agg = aggregate(records)

            for cat, v in agg.items():
                for k in ("Total", "Sold", "Balance", "BalanceValue", "Rows"):
                    combined_agg[cat][k] += v[k]

            df = agg_to_dataframe(agg)
            st.subheader(sheet_name)
            st.dataframe(format_display(df), use_container_width=True, hide_index=True)

            if debug_mode:
                st.markdown("**Debug: row-level detail feeding each category**")
                cat_options = sorted({r["Category"] for r in records})
                chosen_cat = st.selectbox(
                    "Show raw rows contributing to category:",
                    cat_options, key=f"debug_{sheet_name}"
                )
                detail_rows = [r for r in records if r["Category"] == chosen_cat]
                detail_df = pd.DataFrame(detail_rows)[
                    ["Row", "Identifier", "LotTypeRaw", "Total", "Balance", "Sold", "BalanceValue"]
                ]
                st.dataframe(detail_df, use_container_width=True, hide_index=True)
                st.caption(f"{len(detail_rows)} row(s) → Total = {detail_df['Total'].sum():,.0f}")

    with tabs[-1]:
        st.subheader("Combined Summary — All Sheets")
        combined_df = agg_to_dataframe(combined_agg)
        st.dataframe(format_display(combined_df), use_container_width=True, hide_index=True)

        grand_total = combined_df.loc[combined_df["Category"] == "Total", "Total"].values
        if len(grand_total):
            st.metric("Grand Total Units", f"{grand_total[0]:,.0f}")

else:
    st.info("👆 Upload an .xlsx file to get started.")
